#!/usr/bin/env python3
"""
將 dse_5_aggressive_p1p2p3.json 匯出為 Excel，含 Path 1 完整數據。

Path 1 數據來源：若 JSON 無 p1_* 則從 terminal log 補入（2026-03-08 run）。
"""

import json
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    print("請先安裝 openpyxl: pip install openpyxl")
    exit(1)

WORKSPACE = Path(__file__).parent
JSON_PATH = WORKSPACE / "dse_5_aggressive_p1p2p3.json"
OUT_PATH = WORKSPACE / "dse_5_aggressive_results.xlsx"

# Path 1 數據（來自 2026-03-08 terminal log）
P1_FROM_LOG = {
    1: {"accuracy": 0.9364, "energy_uj": 0.022, "timing_us": 0.062, "area_mm2": 0.1426},
    2: {"accuracy": 0.7688, "energy_uj": 0.022, "timing_us": 0.062, "area_mm2": 0.1426},
    3: {"accuracy": 0.9391, "energy_uj": 0.022, "timing_us": 0.062, "area_mm2": 0.1426},
    4: {"accuracy": 0.9291, "energy_uj": 0.022, "timing_us": 0.062, "area_mm2": 0.1426},
    5: {"accuracy": 0.9446, "energy_uj": 0.022, "timing_us": 0.062, "area_mm2": 0.1426},
}


def main():
    with open(JSON_PATH, encoding="utf-8") as f:
        data = json.load(f)

    results = data["results"]

    # 補入 Path 1 數據（若 JSON 無則用 terminal log）
    for r in results:
        dp = r.get("dp")
        if dp in P1_FROM_LOG and "p1_accuracy" not in r:
            p1 = P1_FROM_LOG[dp]
            r["p1_accuracy"] = p1["accuracy"]
            r["p1_energy_uj"] = p1["energy_uj"]
            r["p1_timing_us"] = p1["timing_us"]
            r["p1_area_mm2"] = p1["area_mm2"]
        elif "p1_accuracy" not in r:
            r["p1_accuracy"] = r.get("accuracy")
            r["p1_energy_uj"] = r.get("p1_energy_uj")
            r["p1_timing_us"] = r.get("p1_timing_us")
            r["p1_area_mm2"] = r.get("p1_area_mm2")

    # 回寫 JSON 以持久化 Path 1 數據
    with open(JSON_PATH, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2)

    wb = Workbook()
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    ws = wb.active
    ws.title = "5-Point Aggressive (200-300MHz)"

    headers = [
        "DP", "組別", "頻率(MHz)", "hd_dim", "reram", "CNN(oc×k)", "inner_dim",
        "P1_Accuracy(%)", "P1_Energy(µJ)", "P1_Timing(µs)", "P1_Area(mm²)", "P1_elapsed(s)",
        "Final_Accuracy(%)", "Final_Energy(µJ)", "Final_Timing(µs)", "Final_Area(mm²)",
        "p2_slack_ns", "p3_cycles", "p2_elapsed(s)", "p3_elapsed(s)",
    ]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = header_fill
        c.font = header_font

    for row, r in enumerate(results, 2):
        p = r.get("params", {})
        freq = p.get("frequency", 0) / 1e6
        cnn = f"{p.get('out_channels_1',0)}×{p.get('kernel_size_1',0)}, {p.get('out_channels_2',0)}×{p.get('kernel_size_2',0)}"
        ws.cell(row=row, column=1, value=r.get("dp"))
        ws.cell(row=row, column=2, value=r.get("group", ""))
        ws.cell(row=row, column=3, value=round(freq, 0))
        ws.cell(row=row, column=4, value=p.get("hd_dim"))
        ws.cell(row=row, column=5, value=p.get("reram_size"))
        ws.cell(row=row, column=6, value=cnn)
        ws.cell(row=row, column=7, value=p.get("inner_dim"))
        p1_acc = r.get("p1_accuracy") or r.get("accuracy")
        ws.cell(row=row, column=8, value=round((p1_acc or 0) * 100, 2) if p1_acc is not None else None)
        ws.cell(row=row, column=9, value=round(r.get("p1_energy_uj", 0), 4) if r.get("p1_energy_uj") is not None else None)
        ws.cell(row=row, column=10, value=round(r.get("p1_timing_us", 0), 4) if r.get("p1_timing_us") is not None else None)
        ws.cell(row=row, column=11, value=round(r.get("p1_area_mm2", 0), 4) if r.get("p1_area_mm2") is not None else None)
        ws.cell(row=row, column=12, value=r.get("p1_elapsed_s"))
        ws.cell(row=row, column=13, value=round(r.get("accuracy", 0) * 100, 2))
        ws.cell(row=row, column=14, value=round(r.get("energy_uj", 0), 2))
        ws.cell(row=row, column=15, value=round(r.get("timing_us", 0), 2))
        ws.cell(row=row, column=16, value=round(r.get("area_mm2", 0), 2))
        ws.cell(row=row, column=17, value=r.get("p2_timing_slack_ns"))
        ws.cell(row=row, column=18, value=r.get("p3_execution_cycles"))
        ws.cell(row=row, column=19, value=r.get("p2_elapsed_s"))
        ws.cell(row=row, column=20, value=r.get("p3_elapsed_s"))

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 14

    try:
        wb.save(OUT_PATH)
        print(f"已匯出: {OUT_PATH}")
    except PermissionError:
        alt = WORKSPACE / "dse_5_aggressive_results_new.xlsx"
        wb.save(alt)
        print(f"主檔被佔用，已匯出至: {alt}")


if __name__ == "__main__":
    main()
