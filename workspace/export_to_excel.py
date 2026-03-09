#!/usr/bin/env python3
"""
將 dse_merged_p1p2p3.json 與分析圖表匯出為 Excel 檔。
"""

import json
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    print("請先安裝 openpyxl: pip install openpyxl")
    exit(1)

WORKSPACE = Path(__file__).parent
JSON_PATH = WORKSPACE / "dse_merged_p1p2p3.json"
OUT_PATH = WORKSPACE / "experiment_results_and_analysis.xlsx"
OUT_PATH_ALT = WORKSPACE / "experiment_results_and_analysis_new.xlsx"  # 若主檔被佔用則用此


def _opt_desc(p):
    """產生主要優化設定簡述"""
    parts = []
    if p.get("syn_map_effort") and p.get("syn_opt_effort"):
        parts.append(f"syn:{p['syn_map_effort']}, opt:{p['syn_opt_effort']}")
    if str(p.get("enable_clock_gating", "")).lower() == "true":
        parts.append("cg")
    if str(p.get("enable_retime", "")).lower() == "true":
        parts.append("retime")
    if str(p.get("compile_timing_high_effort", "")).lower() == "true":
        parts.append("timing_high")
    if str(p.get("compile_ultra_gate_clock", "")).lower() == "true":
        parts.append("ultra_gate")
    if str(p.get("enable_leakage_optimization", "")).lower() == "true":
        parts.append("leakage")
    if str(p.get("enable_dynamic_optimization", "")).lower() == "true":
        parts.append("dynamic")
    if str(p.get("max_area_ignore_tns", "")).lower() == "true":
        parts.append("max_area_ignore_tns")
    dp = str(p.get("dp_smartgen_strategy", "none")).lower()
    if dp != "none":
        parts.append(f"dp:{dp}")
    return ", ".join(parts) if parts else "baseline"


def main():
    with open(JSON_PATH, encoding="utf-8") as f:
        data = json.load(f)

    results = data["results"]

    wb = Workbook()
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    # ========== Sheet 1: 實驗結果與設定對照表（含 P1 完整數據） ==========
    ws1 = wb.active
    ws1.title = "實驗結果與設定對照"

    headers1 = [
        "DP", "組別", "頻率(MHz)", "hd_dim", "reram", "CNN(oc×k)", "inner_dim",
        "主要優化設定",
        "P1_Accuracy(%)", "P1_Energy(µJ)", "P1_Timing(µs)", "P1_Area(mm²)", "P1_elapsed(s)",
        "Final_Accuracy(%)", "Final_Energy(µJ)", "Final_Timing(µs)", "Final_Area(mm²)",
        "p2_slack_ns", "p3_cycles", "p2_elapsed(s)", "p3_elapsed(s)",
    ]
    for col, h in enumerate(headers1, 1):
        c = ws1.cell(row=1, column=col, value=h)
        c.fill = header_fill
        c.font = header_font

    for row, r in enumerate(results, 2):
        p = r.get("params", {})
        freq = p.get("frequency", 0) / 1e6
        cnn = f"{p.get('out_channels_1',0)}×{p.get('kernel_size_1',0)}, {p.get('out_channels_2',0)}×{p.get('kernel_size_2',0)}"
        ws1.cell(row=row, column=1, value=r.get("dp"))
        ws1.cell(row=row, column=2, value=r.get("group", ""))
        ws1.cell(row=row, column=3, value=round(freq, 0))
        ws1.cell(row=row, column=4, value=p.get("hd_dim"))
        ws1.cell(row=row, column=5, value=p.get("reram_size"))
        ws1.cell(row=row, column=6, value=cnn)
        ws1.cell(row=row, column=7, value=p.get("inner_dim"))
        ws1.cell(row=row, column=8, value=_opt_desc(p))
        # P1 數據（舊資料無 p1_* 時，accuracy 來自 P1，其餘留空）
        p1_acc = r.get("p1_accuracy") or r.get("accuracy")
        ws1.cell(row=row, column=9, value=round((p1_acc or 0) * 100, 2) if p1_acc is not None else None)
        ws1.cell(row=row, column=10, value=round(r.get("p1_energy_uj", 0), 2) if r.get("p1_energy_uj") is not None else None)
        ws1.cell(row=row, column=11, value=round(r.get("p1_timing_us", 0), 2) if r.get("p1_timing_us") is not None else None)
        ws1.cell(row=row, column=12, value=round(r.get("p1_area_mm2", 0), 2) if r.get("p1_area_mm2") is not None else None)
        ws1.cell(row=row, column=13, value=r.get("p1_elapsed_s"))
        # Final 數據（Path 3 或 Path 2 縫合結果）
        ws1.cell(row=row, column=14, value=round(r.get("accuracy", 0) * 100, 2))
        ws1.cell(row=row, column=15, value=round(r.get("energy_uj", 0), 2))
        ws1.cell(row=row, column=16, value=round(r.get("timing_us", 0), 2))
        ws1.cell(row=row, column=17, value=round(r.get("area_mm2", 0), 2))
        ws1.cell(row=row, column=18, value=r.get("p2_timing_slack_ns"))
        ws1.cell(row=row, column=19, value=r.get("p3_execution_cycles"))
        ws1.cell(row=row, column=20, value=r.get("p2_elapsed_s"))
        ws1.cell(row=row, column=21, value=r.get("p3_elapsed_s"))

    for col in range(1, len(headers1) + 1):
        ws1.column_dimensions[get_column_letter(col)].width = 14

    # ========== Sheet 2: 縮寫說明 ==========
    ws2 = wb.create_sheet("縮寫說明")
    abbrev = [
        ("縮寫", "全名"),
        ("syn", "syn_map_effort / syn_opt_effort"),
        ("cg", "enable_clock_gating"),
        ("retime", "enable_retime"),
        ("timing_high", "compile_timing_high_effort"),
        ("ultra_gate", "compile_ultra_gate_clock"),
        ("leakage", "enable_leakage_optimization"),
        ("dynamic", "enable_dynamic_optimization"),
        ("dp:area", "dp_smartgen_strategy: area"),
        ("dp:timing", "dp_smartgen_strategy: timing"),
        ("CNN (ch×k)", "out_channels_1×kernel_size_1, out_channels_2×kernel_size_2"),
    ]
    for row, (k, v) in enumerate(abbrev, 1):
        ws2.cell(row=row, column=1, value=k)
        ws2.cell(row=row, column=2, value=v)
        if row == 1:
            ws2.cell(row=row, column=1).fill = header_fill
            ws2.cell(row=row, column=1).font = header_font
            ws2.cell(row=row, column=2).fill = header_fill
            ws2.cell(row=row, column=2).font = header_font
    ws2.column_dimensions["A"].width = 18
    ws2.column_dimensions["B"].width = 50

    # ========== Sheet 3: 組別與實驗重點 ==========
    ws3 = wb.create_sheet("組別與實驗重點")
    groups = [
        ("組別", "實驗重點"),
        ("A", "合成/優化策略（syn effort、clock gating、retime、dp_smartgen 等）"),
        ("B", "架構規模（hd_dim、reram_size、CNN 通道與 kernel）"),
        ("C", "inner_dim（1024 vs 4096）"),
        ("D", "頻率（125/150 MHz）與合成優化"),
        ("E", "hd_dim=4096、reram=256 下之頻率與優化"),
        ("G", "頻率掃描（80–175 MHz）與 inner_dim（1024/2048/4096）"),
        ("H", "高頻掃描（200–300 MHz）"),
    ]
    for row, (g, d) in enumerate(groups, 1):
        ws3.cell(row=row, column=1, value=g)
        ws3.cell(row=row, column=2, value=d)
        if row == 1:
            ws3.cell(row=row, column=1).fill = header_fill
            ws3.cell(row=row, column=1).font = header_font
            ws3.cell(row=row, column=2).fill = header_fill
            ws3.cell(row=row, column=2).font = header_font
    ws3.column_dimensions["A"].width = 10
    ws3.column_dimensions["B"].width = 55

    # ========== Sheet 4: A組 Accuracy 差異分析 ==========
    ws4 = wb.create_sheet("A組Accuracy分析")
    a_group = [r for r in results if r.get("group") == "A"]
    headers4 = ["DP", "策略", "Accuracy", "Energy(µJ)", "Timing(µs)", "備註"]
    for col, h in enumerate(headers4, 1):
        c = ws4.cell(row=1, column=col, value=h)
        c.fill = header_fill
        c.font = header_font

    strategies = {
        1: "low/low 預設",
        2: "high/high, retime, timing_high",
        3: "cg, ultra_gate, leakage, dynamic",
        4: "area_aggressive, max_area_ignore_tns, dp=area",
        5: "timing_aggressive (dp=timing)",
    }
    for row, r in enumerate(a_group, 2):
        dp = r.get("dp")
        ws4.cell(row=row, column=1, value=dp)
        ws4.cell(row=row, column=2, value=strategies.get(dp, ""))
        ws4.cell(row=row, column=3, value=round(r.get("accuracy", 0) * 100, 2))
        ws4.cell(row=row, column=4, value=round(r.get("energy_uj", 0), 2))
        ws4.cell(row=row, column=5, value=round(r.get("timing_us", 0), 2))
        note = "最高" if r.get("accuracy", 0) > 0.95 else ("明顯偏低" if r.get("accuracy", 0) < 0.8 else "")
        ws4.cell(row=row, column=6, value=note)
    for col in range(1, 7):
        ws4.column_dimensions[get_column_letter(col)].width = 18

    # ========== Sheet 5: 建議修正（訓練變異） ==========
    ws5 = wb.create_sheet("建議修正")
    fixes = [
        ("項目", "目前", "建議", "說明"),
        ("cnn_epochs", "1", "10", "MNIST 註解建議 10"),
        ("hd_epochs", "1", "10", "MNIST 註解建議 10"),
        ("num_tests", "1", "3–5", "多次 inference 取平均"),
        ("seed", "config 有", "evaluate_path1 前呼叫 set_seed(42)", "確保可重現"),
    ]
    for row, row_data in enumerate(fixes, 1):
        for col, val in enumerate(row_data, 1):
            c = ws5.cell(row=row, column=col, value=val)
            if row == 1:
                c.fill = header_fill
                c.font = header_font
    ws5.column_dimensions["A"].width = 14
    ws5.column_dimensions["B"].width = 18
    ws5.column_dimensions["C"].width = 28
    ws5.column_dimensions["D"].width = 28

    # ========== Sheet 6: 完整參數（扁平化） ==========
    ws6 = wb.create_sheet("完整參數")
    if results:
        all_keys = set()
        for r in results:
            all_keys.update(r.get("params", {}).keys())
        all_keys = sorted(all_keys)
        headers6 = ["dp", "group", "status", "accuracy", "energy_uj", "timing_us", "area_mm2"] + list(all_keys)
        for col, h in enumerate(headers6, 1):
            c = ws6.cell(row=1, column=col, value=h)
            c.fill = header_fill
            c.font = header_font

        for row, r in enumerate(results, 2):
            ws6.cell(row=row, column=1, value=r.get("dp"))
            ws6.cell(row=row, column=2, value=r.get("group"))
            ws6.cell(row=row, column=3, value=r.get("status"))
            ws6.cell(row=row, column=4, value=r.get("accuracy"))
            ws6.cell(row=row, column=5, value=r.get("energy_uj"))
            ws6.cell(row=row, column=6, value=r.get("timing_us"))
            ws6.cell(row=row, column=7, value=r.get("area_mm2"))
            p = r.get("params", {})
            for col, k in enumerate(all_keys, 8):
                ws6.cell(row=row, column=col, value=p.get(k, ""))

    try:
        wb.save(OUT_PATH)
        print(f"已匯出: {OUT_PATH}")
    except PermissionError:
        wb.save(OUT_PATH_ALT)
        print(f"主檔被佔用，已匯出至: {OUT_PATH_ALT}")


if __name__ == "__main__":
    main()
