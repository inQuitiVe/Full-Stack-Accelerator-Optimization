#!/usr/bin/env python3
"""
將 dse_merged_p1p2p3.json 匯出為完整 Excel 檔。
含：DP、組別、所有參數、Path 1/2/3 完整 metrics、meta 摘要。
"""

import json
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    print("請先安裝 openpyxl: pip install openpyxl")
    exit(1)

WORKSPACE = Path(__file__).parent
JSON_PATH = WORKSPACE / "dse_merged_p1p2p3.json"
OUT_PATH = WORKSPACE / "dse_merged_results.xlsx"
OUT_PATH_ALT = WORKSPACE / "dse_merged_results_new.xlsx"


def _opt_desc(p):
    """產生主要優化設定簡述"""
    parts = []
    if p.get("syn_map_effort") and p.get("syn_opt_effort"):
        parts.append(f"syn:{p['syn_map_effort']}, opt:{p['syn_opt_effort']}")
    if str(p.get("enable_clock_gating", "")).lower() == "true":
        parts.append("cg")
    if str(p.get("enable_retime", "")).lower() == "true":
        parts.append("retime")
    if str(p.get("compile_timing_high_effort", "")).lower() == "true":
        parts.append("timing_high")
    if str(p.get("compile_ultra_gate_clock", "")).lower() == "true":
        parts.append("ultra_gate")
    if str(p.get("enable_leakage_optimization", "")).lower() == "true":
        parts.append("leakage")
    if str(p.get("enable_dynamic_optimization", "")).lower() == "true":
        parts.append("dynamic")
    if str(p.get("max_area_ignore_tns", "")).lower() == "true":
        parts.append("max_area_ignore_tns")
    dp = str(p.get("dp_smartgen_strategy", "none")).lower()
    if dp != "none":
        parts.append(f"dp:{dp}")
    return ", ".join(parts) if parts else "baseline"


def main():
    with open(JSON_PATH, encoding="utf-8") as f:
        data = json.load(f)

    results = data["results"]
    meta = data.get("meta", {})

    wb = Workbook()
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    # ========== Sheet 1: 實驗結果與設定對照（精簡摘要） ==========
    ws1 = wb.active
    ws1.title = "實驗結果與設定對照"

    headers1 = [
        "DP", "組別", "頻率(MHz)", "hd_dim", "reram", "CNN(oc×k)", "inner_dim",
        "主要優化設定",
        "P1_Accuracy(%)", "P1_Energy(µJ)", "P1_Timing(µs)", "P1_Area(mm²)", "P1_elapsed(s)",
        "Final_Accuracy(%)", "Final_Energy(µJ)", "Final_Timing(µs)", "Final_Area(mm²)",
        "p2_area_um2", "p2_slack_ns", "p2_power_mw", "p3_cycles", "p3_power_mw",
        "p2_elapsed(s)", "p3_elapsed(s)",
    ]
    for col, h in enumerate(headers1, 1):
        c = ws1.cell(row=1, column=col, value=h)
        c.fill = header_fill
        c.font = header_font

    for row, r in enumerate(results, 2):
        p = r.get("params", {})
        freq = p.get("frequency", 0) / 1e6
        cnn = f"{p.get('out_channels_1',0)}×{p.get('kernel_size_1',0)}, {p.get('out_channels_2',0)}×{p.get('kernel_size_2',0)}"
        ws1.cell(row=row, column=1, value=r.get("dp"))
        ws1.cell(row=row, column=2, value=r.get("group", ""))
        ws1.cell(row=row, column=3, value=round(freq, 0))
        ws1.cell(row=row, column=4, value=p.get("hd_dim"))
        ws1.cell(row=row, column=5, value=p.get("reram_size"))
        ws1.cell(row=row, column=6, value=cnn)
        ws1.cell(row=row, column=7, value=p.get("inner_dim"))
        ws1.cell(row=row, column=8, value=_opt_desc(p))
        p1_acc = r.get("p1_accuracy") or r.get("accuracy")
        ws1.cell(row=row, column=9, value=round((p1_acc or 0) * 100, 2) if p1_acc is not None else None)
        ws1.cell(row=row, column=10, value=round(r.get("p1_energy_uj", 0), 2) if r.get("p1_energy_uj") is not None else None)
        ws1.cell(row=row, column=11, value=round(r.get("p1_timing_us", 0), 2) if r.get("p1_timing_us") is not None else None)
        ws1.cell(row=row, column=12, value=round(r.get("p1_area_mm2", 0), 2) if r.get("p1_area_mm2") is not None else None)
        ws1.cell(row=row, column=13, value=r.get("p1_elapsed_s"))
        ws1.cell(row=row, column=14, value=round(r.get("accuracy", 0) * 100, 2))
        ws1.cell(row=row, column=15, value=round(r.get("energy_uj", 0), 2))
        ws1.cell(row=row, column=16, value=round(r.get("timing_us", 0), 2))
        ws1.cell(row=row, column=17, value=round(r.get("area_mm2", 0), 2))
        ws1.cell(row=row, column=18, value=round(r.get("p2_area_um2", 0), 2) if r.get("p2_area_um2") is not None else None)
        ws1.cell(row=row, column=19, value=r.get("p2_timing_slack_ns"))
        ws1.cell(row=row, column=20, value=round(r.get("p2_dynamic_power_mw", 0), 4) if r.get("p2_dynamic_power_mw") is not None else None)
        ws1.cell(row=row, column=21, value=r.get("p3_execution_cycles"))
        ws1.cell(row=row, column=22, value=round(r.get("p3_dynamic_power_mw", 0), 4) if r.get("p3_dynamic_power_mw") is not None else None)
        ws1.cell(row=row, column=23, value=r.get("p2_elapsed_s"))
        ws1.cell(row=row, column=24, value=r.get("p3_elapsed_s"))

    for col in range(1, len(headers1) + 1):
        ws1.column_dimensions[get_column_letter(col)].width = 14

    # ========== Sheet 2: 完整資料（所有欄位扁平化） ==========
    ws2 = wb.create_sheet("完整資料")

    # 收集所有頂層 keys（排除 params）
    top_keys = set()
    for r in results:
        for k in r.keys():
            if k != "params":
                top_keys.add(k)
    top_keys = sorted(top_keys)

    # 收集所有 params keys
    param_keys = set()
    for r in results:
        param_keys.update(r.get("params", {}).keys())
    param_keys = sorted(param_keys)

    headers2 = ["dp", "group", "status"] + [k for k in top_keys if k not in ("dp", "group", "status")] + list(param_keys)
    for col, h in enumerate(headers2, 1):
        c = ws2.cell(row=1, column=col, value=h)
        c.fill = header_fill
        c.font = header_font

    for row, r in enumerate(results, 2):
        col = 1
        ws2.cell(row=row, column=col, value=r.get("dp"))
        col += 1
        ws2.cell(row=row, column=col, value=r.get("group"))
        col += 1
        ws2.cell(row=row, column=col, value=r.get("status"))
        col += 1
        for k in top_keys:
            if k in ("dp", "group", "status"):
                continue
            v = r.get(k)
            if isinstance(v, float):
                ws2.cell(row=row, column=col, value=round(v, 6) if v is not None else None)
            else:
                ws2.cell(row=row, column=col, value=v)
            col += 1
        p = r.get("params", {})
        for k in param_keys:
            v = p.get(k)
            ws2.cell(row=row, column=col, value=v)
            col += 1

    for col in range(1, len(headers2) + 1):
        ws2.column_dimensions[get_column_letter(col)].width = 16

    # ========== Sheet 3: 組別與實驗重點（對應 merged 群組） ==========
    ws3 = wb.create_sheet("組別與實驗重點")
    groups = [
        ("組別", "實驗重點"),
        ("EDA", "EDA 合成/優化策略（syn effort、clock gating、retime、dp_smartgen 等）"),
        ("ARCH", "架構規模（hd_dim、reram_size、CNN 通道與 kernel）"),
        ("INNER", "inner_dim 變化（1024 vs 2048 vs 4096）"),
        ("ARCH_EXT", "架構擴展與頻率（hd_dim=4096、reram=256、頻率 200–300 MHz）"),
        ("FREQ", "頻率掃描（225–300 MHz，base arch）"),
    ]
    for row, (g, d) in enumerate(groups, 1):
        ws3.cell(row=row, column=1, value=g)
        ws3.cell(row=row, column=2, value=d)
        if row == 1:
            ws3.cell(row=row, column=1).fill = header_fill
            ws3.cell(row=row, column=1).font = header_font
            ws3.cell(row=row, column=2).fill = header_fill
            ws3.cell(row=row, column=2).font = header_font
    ws3.column_dimensions["A"].width = 12
    ws3.column_dimensions["B"].width = 55

    # ========== Sheet 4: 縮寫說明 ==========
    ws4 = wb.create_sheet("縮寫說明")
    abbrev = [
        ("縮寫", "全名"),
        ("syn", "syn_map_effort / syn_opt_effort"),
        ("cg", "enable_clock_gating"),
        ("retime", "enable_retime"),
        ("timing_high", "compile_timing_high_effort"),
        ("ultra_gate", "compile_ultra_gate_clock"),
        ("leakage", "enable_leakage_optimization"),
        ("dynamic", "enable_dynamic_optimization"),
        ("dp:area", "dp_smartgen_strategy: area"),
        ("dp:timing", "dp_smartgen_strategy: timing"),
        ("CNN (ch×k)", "out_channels_1×kernel_size_1, out_channels_2×kernel_size_2"),
    ]
    for row, (k, v) in enumerate(abbrev, 1):
        ws4.cell(row=row, column=1, value=k)
        ws4.cell(row=row, column=2, value=v)
        if row == 1:
            ws4.cell(row=row, column=1).fill = header_fill
            ws4.cell(row=row, column=1).font = header_font
            ws4.cell(row=row, column=2).fill = header_fill
            ws4.cell(row=row, column=2).font = header_font
    ws4.column_dimensions["A"].width = 18
    ws4.column_dimensions["B"].width = 50

    # ========== Sheet 5: Meta 摘要 ==========
    ws5 = wb.create_sheet("Meta 摘要")
    meta_rows = [
        ("項目", "值"),
        ("path2_enabled", meta.get("path2_enabled")),
        ("path3_enabled", meta.get("path3_enabled")),
        ("synth_mode", meta.get("synth_mode")),
        ("top_module", meta.get("top_module")),
        ("total_data_points", meta.get("total_data_points")),
        ("run_type", meta.get("run_type")),
        ("run_start_iso", meta.get("run_start_iso")),
        ("run_end_iso", meta.get("run_end_iso")),
        ("wall_clock_seconds", meta.get("wall_clock_seconds")),
    ]
    for row, (k, v) in enumerate(meta_rows, 1):
        ws5.cell(row=row, column=1, value=k)
        ws5.cell(row=row, column=2, value=v)
        if row == 1:
            ws5.cell(row=row, column=1).fill = header_fill
            ws5.cell(row=row, column=1).font = header_font
            ws5.cell(row=row, column=2).fill = header_fill
            ws5.cell(row=row, column=2).font = header_font
    ws5.column_dimensions["A"].width = 22
    ws5.column_dimensions["B"].width = 50

    try:
        wb.save(OUT_PATH)
        print(f"已匯出: {OUT_PATH}")
    except PermissionError:
        wb.save(OUT_PATH_ALT)
        print(f"主檔被佔用，已匯出至: {OUT_PATH_ALT}")


if __name__ == "__main__":
    main()
